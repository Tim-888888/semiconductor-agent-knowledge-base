"""Generic, review-first standardization for unknown future corpus uploads."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import mimetypes
import zipfile
from collections import Counter
from datetime import UTC, datetime
from pathlib import PurePosixPath
from statistics import fmean
from typing import Any, Protocol

from openpyxl import load_workbook

from semikb.config import Settings
from semikb.contracts.corpus import (
    CorpusFileManifest,
    CorpusFileRelation,
    CorpusFileRole,
    CorpusKind,
    CorpusRelationRule,
    CorpusSidecar,
    CorpusStandardizationError,
    CorpusStandardizationEvent,
    CorpusStandardizationJob,
    CorpusStandardizationMetadata,
    CorpusStandardizationReport,
    CorpusStandardizationStatus,
    CorpusUploadedFile,
    TabularColumnProfile,
    TabularDataProfile,
    TabularSheetProfile,
)
from semikb.contracts.models import ObjectRef
from semikb.rag_ingestion.semikb_adapter import SemikbIngestAdapter
from semikb.storage.corpus_standardization import CorpusStandardizationConflictError
from semikb_ingest import IngestError
from semikb_ingest.models import SourceFormat

_MAX_ARCHIVE_ENTRIES = 10_000
_MAX_ARCHIVE_TOTAL_BYTES = 512 * 1024 * 1024
_MAX_ARCHIVE_MEMBER_BYTES = 100 * 1024 * 1024
_MAX_ARCHIVE_RATIO = 100.0
_DANGEROUS_EXTENSIONS = {
    ".bat",
    ".cmd",
    ".com",
    ".dll",
    ".exe",
    ".hta",
    ".jar",
    ".js",
    ".msi",
    ".ps1",
    ".scr",
    ".sh",
    ".vbs",
}

TRANSIENT_CORPUS_FAILURE_CODES = frozenset(
    {
        "APITIMEOUTERROR",
        "CONNECTIONERROR",
        "INGEST_PARSER_TIMEOUT",
        "INGEST_PARSER_UNAVAILABLE",
        "MINERUERROR",
        "S3ERROR",
        "SERVERSELECTIONTIMEOUTERROR",
        "SOFTTIMELIMITEXCEEDED",
        "TIMEOUTERROR",
        "TIMEOUTEXCEPTION",
    }
)


class CorpusStandardizationStore(Protocol):
    def create_or_get(self, job: CorpusStandardizationJob) -> CorpusStandardizationJob: ...

    def save(self, job: CorpusStandardizationJob) -> CorpusStandardizationJob: ...

    def get(self, job_id: str) -> CorpusStandardizationJob | None: ...

    def list(self) -> list[CorpusStandardizationJob]: ...

    def store_raw(self, **kwargs: Any) -> ObjectRef: ...

    def store_derived(self, **kwargs: Any) -> ObjectRef: ...

    def load_object(self, object_ref: ObjectRef) -> bytes: ...


class CorpusStandardizationService:
    """Creates private, immutable standardization snapshots without publishing RAG data."""

    def __init__(
        self,
        store: CorpusStandardizationStore,
        settings: Settings,
        adapter: SemikbIngestAdapter | None = None,
    ) -> None:
        self._store = store
        self._settings = settings
        self._adapter = adapter or SemikbIngestAdapter(settings)

    def submit(
        self,
        uploads: list[CorpusUploadedFile],
        metadata: CorpusStandardizationMetadata,
        sidecar: CorpusSidecar | None,
        created_by: str,
    ) -> CorpusStandardizationJob:
        if not uploads:
            raise CorpusStandardizationError(
                "CORPUS_EMPTY_UPLOAD",
                "At least one corpus file is required.",
            )
        normalized_uploads = self._normalize_uploads(uploads)
        total_bytes = sum(len(item.content) for item in normalized_uploads)
        if total_bytes > self._settings.max_upload_mib * 1024 * 1024:
            raise CorpusStandardizationError(
                "CORPUS_UPLOAD_LIMIT_EXCEEDED",
                f"Corpus upload exceeds the {self._settings.max_upload_mib} MiB request limit.",
            )
        active_sidecar = sidecar or CorpusSidecar()
        snapshot_hash = self._snapshot_hash(normalized_uploads)
        fingerprint = self._request_fingerprint(metadata, active_sidecar, snapshot_hash)
        idempotency_key = f"{metadata.corpus_id}:{metadata.snapshot_version}"
        refs = [
            self._store.store_raw(
                corpus_id=metadata.corpus_id,
                snapshot_hash=snapshot_hash,
                category="sources",
                relative_path=item.relative_path,
                content=item.content,
                content_type=item.content_type,
            )
            for item in normalized_uploads
        ]
        job = CorpusStandardizationJob(
            metadata=metadata,
            sidecar=active_sidecar,
            snapshot_hash=snapshot_hash,
            request_fingerprint=fingerprint,
            idempotency_key=idempotency_key,
            source_refs=refs,
            source_paths=[item.relative_path for item in normalized_uploads],
            created_by=created_by,
            events=[
                CorpusStandardizationEvent(
                    status=CorpusStandardizationStatus.QUEUED,
                    message="Corpus standardization job accepted.",
                    progress=0,
                )
            ],
        )
        return self._store.create_or_get(job)

    def get_job(self, job_id: str) -> CorpusStandardizationJob | None:
        return self._store.get(job_id)

    def list_jobs(self) -> list[CorpusStandardizationJob]:
        return self._store.list()

    def prepare_retry(self, job_id: str) -> CorpusStandardizationJob:
        job = self._require_job(job_id)
        if job.status is not CorpusStandardizationStatus.FAILED:
            return job
        job.attempt += 1
        job.status = CorpusStandardizationStatus.QUEUED
        job.progress = 0
        job.error_code = None
        job.safe_error_summary = None
        job.started_at = None
        job.finished_at = None
        job.report = None
        job.report_ref = None
        job.events.append(
            CorpusStandardizationEvent(
                status=CorpusStandardizationStatus.QUEUED,
                message="Corpus standardization retry accepted.",
                progress=0,
                attempt=job.attempt,
            )
        )
        return self._store.save(job)

    def mark_queue_submission_failed(self, job_id: str) -> CorpusStandardizationJob:
        return self._fail(
            self._require_job(job_id),
            "CORPUS_TASK_QUEUE_UNAVAILABLE",
            "Corpus standardization task queue is unavailable. Retry the job later.",
        )

    def process(self, job_id: str) -> CorpusStandardizationJob:
        job = self._require_job(job_id)
        if job.status is CorpusStandardizationStatus.REVIEW_REQUIRED:
            return job
        if job.status is CorpusStandardizationStatus.FAILED:
            return job
        try:
            self._advance(
                job,
                CorpusStandardizationStatus.VALIDATING,
                "Validating immutable snapshot and declarative profile.",
                8,
            )
            sources = [
                (
                    path,
                    self._store.load_object(ref),
                    ref.content_type,
                    ref,
                )
                for path, ref in zip(job.source_paths, job.source_refs, strict=True)
            ]
            self._advance(
                job,
                CorpusStandardizationStatus.SNAPSHOTTING,
                "Verifying private source objects and expanding safe archives.",
                20,
            )
            inventory_inputs = self._expand_sources(job, sources)
            self._advance(
                job,
                CorpusStandardizationStatus.INVENTORYING,
                "Classifying files with the shared parser dispatcher and sidecar rules.",
                38,
            )
            manifests = self._build_inventory(job, inventory_inputs)
            relations = self._resolve_relations(job.sidecar, manifests)
            self._advance(
                job,
                CorpusStandardizationStatus.STANDARDIZING,
                "Creating reviewable document, table, and image representations.",
                55,
            )
            manifests = self._standardize(job, manifests, inventory_inputs)
            report = self._build_report(job, manifests, relations)
            report_bytes = json.dumps(
                report.model_dump(mode="json"),
                ensure_ascii=False,
                indent=2,
            ).encode("utf-8")
            report_ref = self._store.store_derived(
                corpus_id=job.metadata.corpus_id,
                snapshot_hash=job.snapshot_hash,
                category="reports",
                relative_path="standardization-report.json",
                content=report_bytes,
                content_type="application/json",
            )
            job.report = report
            job.report_ref = report_ref
            job.files_count = len(manifests)
            job.documents_count = sum(item.role is CorpusFileRole.DOCUMENT for item in manifests)
            job.tables_count = sum(item.role in {CorpusFileRole.TABLE, CorpusFileRole.LABEL} for item in manifests)
            job.images_count = sum(item.role is CorpusFileRole.IMAGE for item in manifests)
            job.unsupported_count = sum(item.role is CorpusFileRole.UNSUPPORTED for item in manifests)
            job.status = CorpusStandardizationStatus.REVIEW_REQUIRED
            job.progress = 100
            job.finished_at = datetime.now(UTC)
            job.events.append(
                CorpusStandardizationEvent(
                    status=CorpusStandardizationStatus.REVIEW_REQUIRED,
                    message="Standardized artifacts are ready for human review; nothing was published.",
                    progress=100,
                    attempt=job.attempt,
                )
            )
            return self._store.save(job)
        except CorpusStandardizationError as exc:
            return self._fail(job, exc.code, exc.safe_message)
        except IngestError as exc:
            return self._fail(job, exc.code.value, exc.safe_message)
        except CorpusStandardizationConflictError:
            raise
        except Exception as exc:
            failure_code = type(exc).__name__.upper()
            if failure_code not in TRANSIENT_CORPUS_FAILURE_CODES:
                failure_code = "CORPUS_STANDARDIZATION_FAILED"
            return self._fail(
                job,
                failure_code,
                "Corpus standardization failed safely. No knowledge data was published.",
            )

    def _expand_sources(
        self,
        job: CorpusStandardizationJob,
        sources: list[tuple[str, bytes, str, ObjectRef]],
    ) -> dict[str, tuple[bytes, str, ObjectRef, bool]]:
        expanded: dict[str, tuple[bytes, str, ObjectRef, bool]] = {}
        casefold_paths: set[str] = set()
        for path, content, content_type, ref in sources:
            self._assert_unique_path(path, casefold_paths)
            if PurePosixPath(path).suffix.lower() != ".zip":
                self._reject_dangerous_path(path)
                expanded[path] = (content, content_type, ref, False)
                continue
            expanded[path] = (content, content_type, ref, True)
            archive_prefix = PurePosixPath(path).stem
            try:
                with zipfile.ZipFile(io.BytesIO(content)) as archive:
                    infos = [info for info in archive.infolist() if not info.is_dir()]
                    if len(infos) > _MAX_ARCHIVE_ENTRIES:
                        self._raise_archive_limit()
                    total = 0
                    for info in infos:
                        member = self._normalize_path(info.filename)
                        member_path = self._normalize_path(f"{archive_prefix}/{member}")
                        self._assert_unique_path(member_path, casefold_paths)
                        self._reject_dangerous_path(member_path)
                        if PurePosixPath(member_path).suffix.lower() == ".zip":
                            raise CorpusStandardizationError(
                                "CORPUS_NESTED_ARCHIVE_REJECTED",
                                "Nested ZIP archives are not accepted in one standardization pass.",
                            )
                        unix_mode = (info.external_attr >> 16) & 0o170000
                        if unix_mode == 0o120000:
                            raise CorpusStandardizationError(
                                "CORPUS_ARCHIVE_SYMLINK_REJECTED",
                                "Archive symbolic links are not accepted.",
                            )
                        if info.flag_bits & 0x1:
                            raise CorpusStandardizationError(
                                "CORPUS_ARCHIVE_ENCRYPTED",
                                "Encrypted archive members are not supported.",
                            )
                        if info.file_size > _MAX_ARCHIVE_MEMBER_BYTES:
                            self._raise_archive_limit()
                        total += info.file_size
                        if total > _MAX_ARCHIVE_TOTAL_BYTES:
                            self._raise_archive_limit()
                        if info.file_size:
                            if info.compress_size == 0 or info.file_size / info.compress_size > _MAX_ARCHIVE_RATIO:
                                self._raise_archive_limit()
                        member_content = archive.read(info)
                        member_type = self._media_type_for_path(member_path)
                        member_ref = self._store.store_raw(
                            corpus_id=job.metadata.corpus_id,
                            snapshot_hash=job.snapshot_hash,
                            category="expanded",
                            relative_path=member_path,
                            content=member_content,
                            content_type=member_type,
                        )
                        expanded[member_path] = (member_content, member_type, member_ref, False)
            except CorpusStandardizationError:
                raise
            except zipfile.BadZipFile as exc:
                raise CorpusStandardizationError(
                    "CORPUS_ARCHIVE_CORRUPT",
                    "The uploaded ZIP archive is corrupt.",
                ) from exc
        return expanded

    def _build_inventory(
        self,
        job: CorpusStandardizationJob,
        inputs: dict[str, tuple[bytes, str, ObjectRef, bool]],
    ) -> list[CorpusFileManifest]:
        manifests: list[CorpusFileManifest] = []
        annotations = {self._normalize_path(item.path): item for item in job.sidecar.files}
        unmatched_annotations = set(annotations).difference(inputs)
        if unmatched_annotations:
            raise CorpusStandardizationError(
                "CORPUS_SIDECAR_PATH_NOT_FOUND",
                "Sidecar annotations reference files that are not present in the snapshot.",
            )
        for path in sorted(inputs):
            content, content_type, ref, is_archive = inputs[path]
            if not self._is_included(path, job.sidecar):
                continue
            annotation = annotations.get(path)
            description = annotation.description if annotation else ""
            if is_archive:
                role = CorpusFileRole.ARCHIVE
                source_format = None
                parser_name = None
            else:
                role_override = annotation.role if annotation else None
                rule_roles = {
                    rule.role
                    for rule in job.sidecar.profile.role_rules
                    if PurePosixPath(path).match(rule.glob)
                }
                if role_override is not None:
                    rule_roles.add(role_override)
                if len(rule_roles) > 1:
                    raise CorpusStandardizationError(
                        "CORPUS_AMBIGUOUS_ROLE",
                        "More than one file role matches the same corpus path.",
                    )
                try:
                    route = self._adapter.resolve(path, content, content_type)
                except IngestError as exc:
                    if exc.code.value != "INGEST_UNSUPPORTED_FORMAT":
                        raise
                    route = None
                if route is None:
                    if (
                        annotation is not None
                        and annotation.role in {CorpusFileRole.TABLE, CorpusFileRole.LABEL}
                        and annotation.tabular_delimiter is not None
                    ):
                        role = annotation.role
                        source_format = "delimited_text"
                        parser_name = "tabular-profile-v1"
                    else:
                        role = CorpusFileRole.UNSUPPORTED
                        source_format = None
                        parser_name = None
                else:
                    inferred = self._role_for_format(route.source_format)
                    requested = next(iter(rule_roles), None)
                    if requested is CorpusFileRole.LABEL and inferred is CorpusFileRole.TABLE:
                        role = CorpusFileRole.LABEL
                    elif requested is not None and requested is not inferred:
                        raise CorpusStandardizationError(
                            "CORPUS_ROLE_FORMAT_MISMATCH",
                            "A sidecar role is incompatible with the file format selected by the parser dispatcher.",
                        )
                    else:
                        role = requested or inferred
                    source_format = route.source_format.value
                    parser_name = route.parser_id
            digest = hashlib.sha256(content).hexdigest()
            manifests.append(
                CorpusFileManifest(
                    file_id=self._file_id(path, digest),
                    relative_path=path,
                    sha256=digest,
                    size_bytes=len(content),
                    content_type=content_type,
                    role=role,
                    source_format=source_format,
                    parser_name=parser_name,
                    raw_ref=ref,
                    warning_codes=(
                        ["CORPUS_UNSUPPORTED_FILE_RETAINED"]
                        if role is CorpusFileRole.UNSUPPORTED
                        else []
                    ),
                    description=description,
                    tabular_delimiter=(annotation.tabular_delimiter if annotation else None),
                    tabular_has_header=(annotation.tabular_has_header if annotation else True),
                )
            )
        if not manifests:
            raise CorpusStandardizationError(
                "CORPUS_NO_INCLUDED_FILES",
                "The corpus profile excluded every uploaded file.",
            )
        return manifests

    def _standardize(
        self,
        job: CorpusStandardizationJob,
        manifests: list[CorpusFileManifest],
        inputs: dict[str, tuple[bytes, str, ObjectRef, bool]],
    ) -> list[CorpusFileManifest]:
        results: list[CorpusFileManifest] = []
        for manifest in manifests:
            if manifest.role in {CorpusFileRole.ARCHIVE, CorpusFileRole.UNSUPPORTED}:
                results.append(manifest)
                continue
            content, content_type, _, _ = inputs[manifest.relative_path]
            if manifest.role in {CorpusFileRole.TABLE, CorpusFileRole.LABEL}:
                profile = self._profile_table(
                    manifest.relative_path,
                    content,
                    manifest.source_format or "",
                    sample_rows=job.sidecar.profile.tabular_sample_rows,
                    max_columns=job.sidecar.profile.tabular_max_columns,
                    delimiter=manifest.tabular_delimiter,
                    has_header=manifest.tabular_has_header,
                )
                payload = json.dumps(
                    profile.model_dump(mode="json"),
                    ensure_ascii=False,
                    indent=2,
                ).encode("utf-8")
                derived = self._store.store_derived(
                    corpus_id=job.metadata.corpus_id,
                    snapshot_hash=job.snapshot_hash,
                    category="standardized",
                    relative_path=f"{manifest.file_id}/tabular-profile.json",
                    content=payload,
                    content_type="application/json",
                )
                results.append(
                    manifest.model_copy(
                        update={"standardized_ref": derived, "tabular_profile": profile}
                    )
                )
                continue
            if manifest.role is CorpusFileRole.IMAGE and not job.sidecar.profile.generate_image_text:
                results.append(
                    manifest.model_copy(
                        update={"warning_codes": [*manifest.warning_codes, "CORPUS_IMAGE_TEXT_DISABLED"]}
                    )
                )
                continue
            session = self._adapter.parse(
                manifest.relative_path,
                content,
                correlation_id=job.job_id,
                declared_media_type=content_type,
            )
            try:
                document = session.document
                markdown = document.normalized_markdown.encode("utf-8")
                derived = self._store.store_derived(
                    corpus_id=job.metadata.corpus_id,
                    snapshot_hash=job.snapshot_hash,
                    category="standardized",
                    relative_path=f"{manifest.file_id}/document.md",
                    content=markdown,
                    content_type="text/markdown",
                )
                results.append(
                    manifest.model_copy(
                        update={
                            "standardized_ref": derived,
                            "parser_name": document.provenance.parser_name,
                            "parser_version": document.provenance.parser_version,
                            "warning_codes": [
                                *manifest.warning_codes,
                                *(warning.code for warning in document.warnings),
                            ],
                        }
                    )
                )
            finally:
                session.discard_remaining()
        return results

    def _profile_table(
        self,
        filename: str,
        content: bytes,
        source_format: str,
        *,
        sample_rows: int,
        max_columns: int,
        delimiter: str | None,
        has_header: bool,
    ) -> TabularDataProfile:
        if source_format in {SourceFormat.CSV.value, "delimited_text"}:
            try:
                text = content.decode("utf-8-sig")
            except UnicodeDecodeError as exc:
                raise CorpusStandardizationError(
                    "CORPUS_TABULAR_ENCODING_UNSUPPORTED",
                    "CSV standardization requires UTF-8 text.",
                ) from exc
            if delimiter == "whitespace":
                iterator = (line.split() for line in io.StringIO(text) if line.strip())
            else:
                actual_delimiter = "\t" if delimiter == "tab" else delimiter or ","
                iterator = csv.reader(io.StringIO(text), delimiter=actual_delimiter)
            rows = []
            for index, row in enumerate(iterator):
                if index > sample_rows + (1 if has_header else 0):
                    break
                rows.append(list(row))
            sheet = self._profile_rows(
                "delimited_text" if source_format == "delimited_text" else "csv",
                rows,
                sample_rows=sample_rows,
                max_columns=max_columns,
                has_header=has_header,
            )
            return TabularDataProfile(sheets=[sheet])
        if source_format == SourceFormat.XLSX.value:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            try:
                sheets: list[TabularSheetProfile] = []
                for worksheet in workbook.worksheets:
                    rows = []
                    for index, row in enumerate(worksheet.iter_rows(values_only=True)):
                        if index > sample_rows:
                            break
                        rows.append(list(row))
                    profile = self._profile_rows(
                        worksheet.title,
                        rows,
                        sample_rows=sample_rows,
                        max_columns=max_columns,
                        has_header=True,
                    )
                    if worksheet.max_row and worksheet.max_row - 1 > profile.observed_rows:
                        profile.sample_truncated = True
                    sheets.append(profile)
                return TabularDataProfile(sheets=sheets)
            finally:
                workbook.close()
        raise CorpusStandardizationError(
            "CORPUS_TABULAR_FORMAT_UNSUPPORTED",
            "Only CSV and XLSX files can use the tabular profile route.",
        )

    @staticmethod
    def _profile_rows(
        name: str,
        rows: list[list[Any]],
        *,
        sample_rows: int,
        max_columns: int,
        has_header: bool,
    ) -> TabularSheetProfile:
        if not rows:
            return TabularSheetProfile(name=name)
        width = max((len(row) for row in rows), default=0)
        bounded_width = min(width, max_columns)
        raw_headers = rows[0][:bounded_width] if has_header else []
        headers = (
            [
                str(value).strip()
                if value is not None and str(value).strip()
                else f"column_{index + 1}"
                for index, value in enumerate(raw_headers)
            ]
            if has_header
            else [f"column_{index + 1}" for index in range(bounded_width)]
        )
        seen: Counter[str] = Counter()
        unique_headers = []
        for header in headers:
            seen[header] += 1
            unique_headers.append(header if seen[header] == 1 else f"{header}_{seen[header]}")
        data_rows = rows[1 : sample_rows + 1] if has_header else rows[:sample_rows]
        columns = []
        for index, header in enumerate(unique_headers):
            values = [row[index] if index < len(row) else None for row in data_rows]
            non_empty = [value for value in values if value not in {None, ""}]
            types = Counter(CorpusStandardizationService._value_type(value) for value in non_empty)
            numerics = [float(value) for value in non_empty if CorpusStandardizationService._is_number(value)]
            columns.append(
                TabularColumnProfile(
                    name=header,
                    inferred_types=dict(types),
                    non_empty_count=len(non_empty),
                    empty_count=len(values) - len(non_empty),
                    numeric_min=min(numerics) if numerics else None,
                    numeric_max=max(numerics) if numerics else None,
                    numeric_mean=fmean(numerics) if numerics else None,
                )
            )
        return TabularSheetProfile(
            name=name,
            observed_rows=len(data_rows),
            column_count=width,
            sample_truncated=len(rows) > sample_rows + (1 if has_header else 0),
            columns_truncated=width > max_columns,
            columns=columns,
        )

    def _resolve_relations(
        self,
        sidecar: CorpusSidecar,
        manifests: list[CorpusFileManifest],
    ) -> list[CorpusFileRelation]:
        by_path = {item.relative_path: item for item in manifests}
        rules = [*sidecar.profile.relation_rules, *sidecar.relations]
        relations: list[CorpusFileRelation] = []
        seen: set[tuple[str, str, str]] = set()
        for rule in rules:
            sources = self._match_relation_paths(rule, by_path, from_side=True)
            targets = self._match_relation_paths(rule, by_path, from_side=False)
            if not sources or not targets:
                raise CorpusStandardizationError(
                    "CORPUS_RELATION_TARGET_NOT_FOUND",
                    "A declared corpus relation does not match both source and target files.",
                )
            for source in sources:
                for target in targets:
                    key = (source.file_id, target.file_id, rule.relation_type.value)
                    if key in seen:
                        continue
                    seen.add(key)
                    relations.append(
                        CorpusFileRelation(
                            from_file_id=source.file_id,
                            to_file_id=target.file_id,
                            relation_type=rule.relation_type,
                        )
                    )
        return relations

    @staticmethod
    def _match_relation_paths(
        rule: CorpusRelationRule,
        by_path: dict[str, CorpusFileManifest],
        *,
        from_side: bool,
    ) -> list[CorpusFileManifest]:
        pattern = rule.from_glob if from_side else rule.to_glob
        return [item for path, item in by_path.items() if PurePosixPath(path).match(pattern)]

    def _build_report(
        self,
        job: CorpusStandardizationJob,
        manifests: list[CorpusFileManifest],
        relations: list[CorpusFileRelation],
    ) -> CorpusStandardizationReport:
        warnings = sorted({code for item in manifests for code in item.warning_codes})
        inferred = self._infer_corpus_kind(manifests)
        review_reasons = ["STANDARDIZATION_REVIEW_REQUIRED"]
        if warnings:
            review_reasons.append("WARNINGS_REQUIRE_REVIEW")
        if job.metadata.corpus_kind is not CorpusKind.AUTO and job.metadata.corpus_kind is not inferred:
            warnings.append("CORPUS_DECLARED_KIND_DIFFERS_FROM_INVENTORY")
            review_reasons.append("CORPUS_KIND_MISMATCH_REVIEW")
        return CorpusStandardizationReport(
            corpus_id=job.metadata.corpus_id,
            snapshot_version=job.metadata.snapshot_version,
            snapshot_hash=job.snapshot_hash,
            inferred_corpus_kind=inferred,
            files=manifests,
            relations=relations,
            warning_codes=sorted(set(warnings)),
            review_reasons=review_reasons,
        )

    @staticmethod
    def _infer_corpus_kind(manifests: list[CorpusFileManifest]) -> CorpusKind:
        roles = {
            item.role
            for item in manifests
            if item.role not in {CorpusFileRole.ARCHIVE, CorpusFileRole.UNSUPPORTED}
        }
        if roles and roles.issubset({CorpusFileRole.TABLE, CorpusFileRole.LABEL}):
            return CorpusKind.TABULAR_DATASET
        if roles == {CorpusFileRole.IMAGE}:
            return CorpusKind.IMAGE_CORPUS
        if roles == {CorpusFileRole.DOCUMENT}:
            return CorpusKind.DOCUMENT_COLLECTION
        return CorpusKind.MIXED

    @staticmethod
    def _role_for_format(source_format: SourceFormat) -> CorpusFileRole:
        if source_format in {SourceFormat.CSV, SourceFormat.XLSX}:
            return CorpusFileRole.TABLE
        if source_format is SourceFormat.IMAGE:
            return CorpusFileRole.IMAGE
        return CorpusFileRole.DOCUMENT

    @staticmethod
    def _media_type_for_path(path: str) -> str:
        stable_types = {
            ".csv": "text/csv",
            ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            ".htm": "text/html",
            ".html": "text/html",
            ".jpeg": "image/jpeg",
            ".jpg": "image/jpeg",
            ".markdown": "text/markdown",
            ".md": "text/markdown",
            ".pdf": "application/pdf",
            ".png": "image/png",
            ".pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
            ".txt": "text/plain",
            ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
        suffix = PurePosixPath(path).suffix.lower()
        return stable_types.get(suffix) or mimetypes.guess_type(path)[0] or "application/octet-stream"

    @staticmethod
    def _is_included(path: str, sidecar: CorpusSidecar) -> bool:
        pure = PurePosixPath(path)
        included = any(pure.match(pattern) for pattern in sidecar.profile.include_globs)
        excluded = any(pure.match(pattern) for pattern in sidecar.profile.exclude_globs)
        return included and not excluded

    @staticmethod
    def _normalize_uploads(uploads: list[CorpusUploadedFile]) -> list[CorpusUploadedFile]:
        result = []
        seen: set[str] = set()
        for upload in uploads:
            path = CorpusStandardizationService._normalize_path(upload.relative_path)
            CorpusStandardizationService._assert_unique_path(path, seen)
            CorpusStandardizationService._reject_dangerous_path(path)
            result.append(upload.model_copy(update={"relative_path": path}))
        return sorted(result, key=lambda item: item.relative_path)

    @staticmethod
    def _normalize_path(value: str) -> str:
        normalized = value.replace("\\", "/").strip("/")
        path = PurePosixPath(normalized)
        if (
            not normalized
            or path.is_absolute()
            or ".." in path.parts
            or any(part in {"", "."} for part in path.parts)
            or (path.parts and path.parts[0].endswith(":"))
        ):
            raise CorpusStandardizationError(
                "CORPUS_UNSAFE_PATH",
                "Corpus file paths must be relative and cannot traverse directories.",
            )
        return path.as_posix()

    @staticmethod
    def _assert_unique_path(path: str, seen: set[str]) -> None:
        folded = path.casefold()
        if folded in seen:
            raise CorpusStandardizationError(
                "CORPUS_DUPLICATE_PATH",
                "Corpus file paths must be unique, including case-insensitive comparisons.",
            )
        seen.add(folded)

    @staticmethod
    def _reject_dangerous_path(path: str) -> None:
        if PurePosixPath(path).suffix.lower() in _DANGEROUS_EXTENSIONS:
            raise CorpusStandardizationError(
                "CORPUS_DANGEROUS_FILE_REJECTED",
                "Executable and script files are not accepted in corpus snapshots.",
            )

    @staticmethod
    def _raise_archive_limit() -> None:
        raise CorpusStandardizationError(
            "CORPUS_ARCHIVE_LIMIT_EXCEEDED",
            "The ZIP archive exceeds the safe entry, size, or compression-ratio limit.",
        )

    @staticmethod
    def _snapshot_hash(uploads: list[CorpusUploadedFile]) -> str:
        digest = hashlib.sha256()
        for upload in uploads:
            content_hash = hashlib.sha256(upload.content).hexdigest()
            digest.update(upload.relative_path.encode("utf-8"))
            digest.update(b"\0")
            digest.update(str(len(upload.content)).encode("ascii"))
            digest.update(b"\0")
            digest.update(content_hash.encode("ascii"))
            digest.update(b"\n")
        return digest.hexdigest()

    @staticmethod
    def _request_fingerprint(
        metadata: CorpusStandardizationMetadata,
        sidecar: CorpusSidecar,
        snapshot_hash: str,
    ) -> str:
        payload = {
            "metadata": metadata.model_dump(mode="json"),
            "sidecar": sidecar.model_dump(mode="json"),
            "snapshot_hash": snapshot_hash,
        }
        return hashlib.sha256(
            json.dumps(payload, sort_keys=True, separators=(",", ":")).encode("utf-8")
        ).hexdigest()

    @staticmethod
    def _file_id(path: str, digest: str) -> str:
        identity = hashlib.sha256(f"{path}\0{digest}".encode()).hexdigest()[:24]
        return f"corpus_file_{identity}"

    @staticmethod
    def _is_number(value: Any) -> bool:
        if isinstance(value, bool) or value is None:
            return False
        try:
            number = float(value)
        except (TypeError, ValueError):
            return False
        return math.isfinite(number)

    @staticmethod
    def _value_type(value: Any) -> str:
        if isinstance(value, bool):
            return "boolean"
        if CorpusStandardizationService._is_number(value):
            return "number"
        if isinstance(value, datetime):
            return "datetime"
        return "text"

    def _advance(
        self,
        job: CorpusStandardizationJob,
        status: CorpusStandardizationStatus,
        message: str,
        progress: int,
    ) -> None:
        job.status = status
        job.progress = progress
        if job.started_at is None:
            job.started_at = datetime.now(UTC)
        job.events.append(
            CorpusStandardizationEvent(
                status=status,
                message=message,
                progress=progress,
                attempt=job.attempt,
            )
        )
        self._store.save(job)

    def _fail(
        self,
        job: CorpusStandardizationJob,
        code: str,
        safe_message: str,
    ) -> CorpusStandardizationJob:
        job.status = CorpusStandardizationStatus.FAILED
        job.error_code = code
        job.safe_error_summary = safe_message
        job.finished_at = datetime.now(UTC)
        job.events.append(
            CorpusStandardizationEvent(
                status=CorpusStandardizationStatus.FAILED,
                message=safe_message,
                progress=job.progress,
                attempt=job.attempt,
            )
        )
        return self._store.save(job)

    def _require_job(self, job_id: str) -> CorpusStandardizationJob:
        job = self._store.get(job_id)
        if job is None:
            raise KeyError(job_id)
        return job
