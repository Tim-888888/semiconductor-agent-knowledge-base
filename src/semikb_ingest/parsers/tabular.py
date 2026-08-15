"""Dedicated CSV and XLSX adapters with bounded table-region extraction."""

from __future__ import annotations

import csv
import io
import re
import time
from dataclasses import dataclass
from html import escape

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from semikb_ingest.errors import IngestError, IngestErrorCode
from semikb_ingest.models import ParseWarning, SourceFormat, SourceLocation, TableAssetDraft
from semikb_ingest.parsers.common import (
    complete_document,
    decode_text,
    matrix_to_html,
    matrix_to_markdown,
    normalized_text,
)
from semikb_ingest.parsers.registry import ParseRequest
from semikb_ingest.structure import BlockKind, StructuredBlock


@dataclass(frozen=True, slots=True)
class TabularLimits:
    max_sheets: int = 64
    max_cells: int = 1_000_000
    max_rows: int = 100_000
    max_columns: int = 512


def _consecutive_groups(values: list[int]) -> list[tuple[int, int]]:
    if not values:
        return []
    groups: list[tuple[int, int]] = []
    start = previous = values[0]
    for value in values[1:]:
        if value != previous + 1:
            groups.append((start, previous))
            start = value
        previous = value
    groups.append((start, previous))
    return groups


def _table_regions(sheet: Worksheet) -> list[tuple[int, int, int, int]]:
    populated_rows = [
        row
        for row in range(1, sheet.max_row + 1)
        if any(
            sheet.cell(row, column).value is not None for column in range(1, sheet.max_column + 1)
        )
    ]
    regions: list[tuple[int, int, int, int]] = []
    for row_start, row_end in _consecutive_groups(populated_rows):
        populated_columns = [
            column
            for column in range(1, sheet.max_column + 1)
            if any(
                sheet.cell(row, column).value is not None for row in range(row_start, row_end + 1)
            )
        ]
        for column_start, column_end in _consecutive_groups(populated_columns):
            regions.append((row_start, row_end, column_start, column_end))
    return regions


def _units(headers: tuple[str, ...]) -> dict[str, str]:
    result: dict[str, str] = {}
    for header in headers:
        match = re.search(r"(?:\(([^()]+)\)|（([^（）]+)）|\[([^\[\]]+)\])\s*$", header)
        if match:
            result[header] = next(value for value in match.groups() if value)
    return result


def _worksheet_html(
    sheet: Worksheet,
    rows: list[list[object]],
    bounds: tuple[int, int, int, int],
) -> str:
    row_start, row_end, column_start, column_end = bounds
    merged: dict[tuple[int, int], tuple[int, int]] = {}
    covered: set[tuple[int, int]] = set()
    for cell_range in sheet.merged_cells.ranges:
        if (
            cell_range.max_row < row_start
            or cell_range.min_row > row_end
            or cell_range.max_col < column_start
            or cell_range.min_col > column_end
        ):
            continue
        top_left = (cell_range.min_row, cell_range.min_col)
        merged[top_left] = (
            cell_range.max_row - cell_range.min_row + 1,
            cell_range.max_col - cell_range.min_col + 1,
        )
        for row in range(cell_range.min_row, cell_range.max_row + 1):
            for column in range(cell_range.min_col, cell_range.max_col + 1):
                if (row, column) != top_left:
                    covered.add((row, column))

    rendered_rows: list[str] = []
    for relative_row, values in enumerate(rows):
        absolute_row = row_start + relative_row
        cells: list[str] = []
        tag = "th" if relative_row == 0 else "td"
        for relative_column, value in enumerate(values):
            absolute_column = column_start + relative_column
            if (absolute_row, absolute_column) in covered:
                continue
            span = merged.get((absolute_row, absolute_column))
            attributes = ""
            if span:
                row_span, column_span = span
                attributes = f' rowspan="{row_span}" colspan="{column_span}"'
            cells.append(f"<{tag}{attributes}>{escape(normalized_text(value))}</{tag}>")
        rendered_rows.append("<tr>" + "".join(cells) + "</tr>")
    if not rendered_rows:
        return ""
    return (
        "<table><thead>"
        + rendered_rows[0]
        + "</thead><tbody>"
        + "".join(rendered_rows[1:])
        + "</tbody></table>"
    )


class CsvStructuredParser:
    parser_id = "csv-structured-v1"
    parser_version = "1.0.0"
    source_format = SourceFormat.CSV

    def __init__(self, limits: TabularLimits | None = None) -> None:
        self.limits = limits or TabularLimits()

    def parse(self, request: ParseRequest):
        started = time.monotonic()
        text, fallback_encoding = decode_text(request.content)
        sample = text[:8192]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(io.StringIO(text), dialect)
        rows: list[list[str]] = []
        try:
            for row_index, row in enumerate(reader, start=1):
                if row_index > self.limits.max_rows:
                    self._raise_limit()
                if len(row) > self.limits.max_columns:
                    self._raise_limit()
                rows.append([normalized_text(value) for value in row])
        except csv.Error as exc:
            raise IngestError(
                IngestErrorCode.PARSE_FAILED,
                "The CSV structure could not be parsed safely.",
            ) from exc
        if sum(len(row) for row in rows) > self.limits.max_cells:
            self._raise_limit()
        markdown, headers = matrix_to_markdown(rows)
        location = SourceLocation(row_start=1, row_end=len(rows))
        table = TableAssetDraft(
            asset_id="table_0001",
            title=request.filename,
            html=matrix_to_html(rows),
            markdown=markdown,
            headers=headers,
            row_count=max(len(rows) - 1, 0),
            column_count=max((len(row) for row in rows), default=0),
            location=location,
        )
        block = StructuredBlock(
            block_id="block_0001",
            kind=BlockKind.TABLE,
            text=markdown,
            location=location,
            table_asset_ids=(table.asset_id,),
            metadata={
                "delimiter": getattr(dialect, "delimiter", ","),
                "units": _units(headers),
            },
        )
        warnings = (
            (
                ParseWarning(
                    code="INGEST_WARNING_TEXT_ENCODING_FALLBACK",
                    safe_message=f"Decoded text using {fallback_encoding}.",
                ),
            )
            if fallback_encoding
            else ()
        )
        return complete_document(
            request=request,
            source_format=self.source_format,
            parser_name=self.parser_id,
            parser_version=self.parser_version,
            provider_name=request.route.provider,
            provider_version="python-csv",
            blocks=(block,),
            tables=(table,),
            warnings=warnings,
            normalized_markdown=markdown,
            started_at=started,
        )

    @staticmethod
    def _raise_limit() -> None:
        raise IngestError(
            IngestErrorCode.DOCUMENT_LIMIT_EXCEEDED,
            "The CSV exceeds configured row, column, or cell limits.",
        )


class XlsxStructuredParser:
    parser_id = "xlsx-structured-v1"
    parser_version = "1.0.0"
    source_format = SourceFormat.XLSX

    def __init__(self, limits: TabularLimits | None = None) -> None:
        self.limits = limits or TabularLimits()

    def parse(self, request: ParseRequest):
        started = time.monotonic()
        try:
            formulas = load_workbook(io.BytesIO(request.content), data_only=False, read_only=False)
            cached = load_workbook(io.BytesIO(request.content), data_only=True, read_only=False)
        except Exception as exc:
            raise IngestError(
                IngestErrorCode.CORRUPT_DOCUMENT,
                "The XLSX workbook could not be opened.",
            ) from exc
        if len(formulas.worksheets) > self.limits.max_sheets:
            self._raise_limit()

        blocks: list[StructuredBlock] = []
        tables: list[TableAssetDraft] = []
        warnings: list[ParseWarning] = []
        total_cells = 0
        hidden_count = 0
        formula_without_cache = 0
        visible_sheets = 0
        for sheet in formulas.worksheets:
            if sheet.sheet_state != "visible":
                hidden_count += 1
                continue
            visible_sheets += 1
            cached_sheet = cached[sheet.title]
            for row_start, row_end, column_start, column_end in _table_regions(sheet):
                row_count = row_end - row_start + 1
                column_count = column_end - column_start + 1
                total_cells += row_count * column_count
                if (
                    row_count > self.limits.max_rows
                    or column_count > self.limits.max_columns
                    or total_cells > self.limits.max_cells
                ):
                    self._raise_limit()
                rows: list[list[object]] = []
                for row in range(row_start, row_end + 1):
                    values: list[object] = []
                    for column in range(column_start, column_end + 1):
                        source_cell: Cell = sheet.cell(row, column)
                        value = source_cell.value
                        if source_cell.data_type == "f":
                            cached_value = cached_sheet.cell(row, column).value
                            if cached_value is not None:
                                value = cached_value
                            else:
                                formula_without_cache += 1
                        values.append(value)
                    rows.append(values)
                markdown, headers = matrix_to_markdown(rows)
                if not markdown:
                    continue
                coordinate = (
                    f"{sheet.cell(row_start, column_start).coordinate}:"
                    f"{sheet.cell(row_end, column_end).coordinate}"
                )
                location = SourceLocation(
                    sheet_name=sheet.title,
                    cell_range=coordinate,
                    row_start=row_start,
                    row_end=row_end,
                )
                asset_id = f"table_{len(tables) + 1:04d}"
                merged_ranges = [
                    str(cell_range)
                    for cell_range in sheet.merged_cells.ranges
                    if not (
                        cell_range.max_row < row_start
                        or cell_range.min_row > row_end
                        or cell_range.max_col < column_start
                        or cell_range.min_col > column_end
                    )
                ]
                table = TableAssetDraft(
                    asset_id=asset_id,
                    title=f"{sheet.title}!{coordinate}",
                    html=_worksheet_html(
                        sheet, rows, (row_start, row_end, column_start, column_end)
                    ),
                    markdown=markdown,
                    headers=headers,
                    row_count=max(row_count - 1, 0),
                    column_count=column_count,
                    location=location,
                )
                tables.append(table)
                blocks.append(
                    StructuredBlock(
                        block_id=f"block_{len(blocks) + 1:04d}",
                        kind=BlockKind.TABLE,
                        text=markdown,
                        heading_path=(sheet.title,),
                        location=location,
                        table_asset_ids=(asset_id,),
                        metadata={
                            "merged_ranges": merged_ranges,
                            "units": _units(headers),
                        },
                    )
                )
        formulas.close()
        cached.close()

        if hidden_count:
            warnings.append(
                ParseWarning(
                    code="INGEST_WARNING_HIDDEN_SHEETS_SKIPPED",
                    safe_message=f"Skipped {hidden_count} hidden worksheet(s).",
                )
            )
        if formula_without_cache:
            warnings.append(
                ParseWarning(
                    code="INGEST_WARNING_FORMULA_CACHE_MISSING",
                    safe_message=(
                        f"Retained {formula_without_cache} formula expression(s) because cached "
                        "values were unavailable; formulas were not executed."
                    ),
                )
            )
        normalized = "\n\n".join(f"## {block.heading_path[0]}\n\n{block.text}" for block in blocks)
        return complete_document(
            request=request,
            source_format=self.source_format,
            parser_name=self.parser_id,
            parser_version=self.parser_version,
            provider_name=request.route.provider,
            provider_version="openpyxl-3.1",
            blocks=blocks,
            tables=tables,
            warnings=warnings,
            detected_title=request.filename.rsplit(".", 1)[0],
            normalized_markdown=normalized,
            sheets=visible_sheets,
            started_at=started,
            reference_knowhere=True,
        )

    @staticmethod
    def _raise_limit() -> None:
        raise IngestError(
            IngestErrorCode.DOCUMENT_LIMIT_EXCEEDED,
            "The XLSX workbook exceeds configured sheet, row, column, or cell limits.",
        )
